"""
app.py — Musca domestica Cohort Simulation  (Streamlit)
─────────────────────────────────────────────────────────────────────────────
Interactive three-temperature life-table simulation for Musca domestica.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

Deploy (Streamlit Community Cloud — free):
    1. Push this folder to a public GitHub repository
    2. Go to https://share.streamlit.io → "New app" → select your repo
    3. Set main file path to  app.py

Embed in a website (after deployment):
    <iframe
      src="https://YOUR-USERNAME-musca-sim.streamlit.app"
      width="100%" height="900px"
      style="border:none; border-radius:8px;">
    </iframe>
─────────────────────────────────────────────────────────────────────────────
"""

import io
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from simulation import (
    DEFAULTS, COLORS, LABELS,
    make_params, run_simulation, build_life_table, calc_demog,
)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Musca domestica Simulation",
    page_icon="🪰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Minimal CSS tweaks ─────────────────────────────────────────────────────
st.markdown("""
<style>
  .metric-card {
    background: #f8fafc;
    border-radius: 10px;
    border-left: 5px solid var(--card-color, #1D4ED8);
    padding: 0.9rem 1.1rem;
    margin-bottom: 0.5rem;
  }
  .metric-card h4 { margin: 0 0 0.5rem 0; font-size: 1.05rem; }
  .metric-card table { width: 100%; font-size: 0.88rem; border-collapse: collapse; }
  .metric-card tr:nth-child(even) { background: #f1f5f9; }
  .metric-card td { padding: 3px 6px; }
  .metric-card td:last-child { text-align: right; font-weight: 600; }
</style>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════════════
# SIDEBAR — controls
# ═════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.title("⚙️ Simulation settings")
    st.markdown("---")

    # Temperature selection
    st.subheader("Temperatures")
    sel_15 = st.checkbox("15 °C", value=True)
    sel_20 = st.checkbox("20 °C", value=True)
    sel_25 = st.checkbox("25 °C", value=True)
    temps_sel = [t for t, on in [("15", sel_15), ("20", sel_20), ("25", sel_25)] if on]

    st.markdown("---")

    # Global simulation settings
    sim_days = st.slider("Simulation length (days)", 30, 150, 110, step=5)
    seed     = st.number_input("Random seed", min_value=1, max_value=99999, value=42, step=1)

    st.markdown("---")

    # Per-temperature parameter expanders
    st.subheader("Biological parameters")

    user_params = {}
    for t in ["15", "20", "25"]:
        d   = DEFAULTS[t]
        col = COLORS[t]
        with st.expander(f"**{LABELS[t]} parameters**", expanded=False):
            st.markdown(f"<span style='color:{col};font-weight:700'>Stage survival (per-day probability)</span>",
                        unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            egg_surv   = c1.slider("Egg",   0.50, 1.00, d["egg_surv"],   0.005, key=f"es_{t}")
            larva_surv = c2.slider("Larva", 0.50, 1.00, d["larva_surv"], 0.005, key=f"ls_{t}")
            pupa_surv  = c3.slider("Pupa",  0.50, 1.00, d["pupa_surv"],  0.005, key=f"ps_{t}")

            st.markdown(f"<span style='color:{col};font-weight:700'>Adult lifespan (days)</span>",
                        unsafe_allow_html=True)
            ca, cb = st.columns(2)
            adult_mean = ca.slider("Mean", 5,   120, d["adult_mean"], 1, key=f"am_{t}")
            adult_sd   = cb.slider("SD",   1.0, 30.0, float(d["adult_sd"]), 0.5, key=f"asd_{t}")

            st.markdown(f"<span style='color:{col};font-weight:700'>Fecundity</span>",
                        unsafe_allow_html=True)
            cf, cg, ch = st.columns(3)
            pre_repro  = cf.slider("Pre-repro (d)", 1,   30,   d["pre_repro"],        1,   key=f"pr_{t}")
            peak_eggs  = cg.slider("Peak eggs/♀/d", 1.0, 60.0, float(d["peak_eggs"]), 0.5, key=f"pe_{t}")
            fec_sigma  = ch.slider("Spread σ",      1.0, 20.0, float(d["fec_sigma"]), 0.5, key=f"fs_{t}")

        user_params[t] = dict(
            adult_mean=adult_mean, adult_sd=adult_sd,
            adult_min=d["adult_min"], adult_max=d["adult_max"],
            egg_surv=egg_surv, larva_surv=larva_surv, pupa_surv=pupa_surv,
            pre_repro=pre_repro, peak_day=d["peak_day"],
            fec_sigma=fec_sigma, peak_eggs=peak_eggs,
        )

    st.markdown("---")

    # Run button
    run_btn = st.button("▶  Run simulation", type="primary", use_container_width=True)


# ═════════════════════════════════════════════════════════════════════════════
# RUN SIMULATION — triggered by button, cached in session state
# ═════════════════════════════════════════════════════════════════════════════
if run_btn:
    if not temps_sel:
        st.warning("Please select at least one temperature.")
        st.stop()

    np.random.seed(int(seed))
    results = {}
    with st.spinner("Running simulation…"):
        for t in temps_sel:
            p   = make_params(int(t), user_params[t])
            sim = run_simulation(p, int(sim_days))
            lt  = build_life_table(sim)
            dem = calc_demog(lt, sim)
            results[t] = dict(p=p, sim=sim, lt=lt, dem=dem)
    st.session_state["results"]  = results
    st.session_state["sim_days"] = sim_days

# Load from session state
results  = st.session_state.get("results", {})
sim_days = st.session_state.get("sim_days", sim_days)


# ═════════════════════════════════════════════════════════════════════════════
# HEADER
# ═════════════════════════════════════════════════════════════════════════════
st.title("🪰 Musca domestica  Cohort Simulation")
st.caption(
    "Three-temperature interactive life-table model  •  "
    "Krebs (2014) Ch. 8  •  Henderson (2016) §12.4  •  "
    "Flint *et al.* (2025)  •  Ali *et al.* (2024)"
)

if not results:
    st.info("👈 Adjust parameters in the sidebar and press **Run simulation** to begin.")
    st.stop()


# ═════════════════════════════════════════════════════════════════════════════
# TABS
# ═════════════════════════════════════════════════════════════════════════════
tab_plot, tab_lt, tab_dem, tab_about = st.tabs(
    ["📈 Population plots", "📊 Life table", "🧮 Demographics", "ℹ️ About"]
)


# ── Tab 1: Six-panel figure ───────────────────────────────────────────────
with tab_plot:
    fig = plt.figure(figsize=(14, 8))
    fig.patch.set_facecolor("white")
    gs  = gridspec.GridSpec(2, 3, hspace=0.42, wspace=0.32,
                            left=0.07, right=0.97, top=0.93, bottom=0.09)

    panel_cfg = [
        ("Surviving females (N\u2093)",        "Age (days)", "N females",       "n_fem",  False),
        ("Age-specific fecundity (m\u2093)",   "Age (days)", "Eggs / \u2640 / day", "mx",  False),
        ("Daily eggs collected",               "Day",        "Eggs",            "eggs",   False),
        ("Survivorship (l\u2093)",             "Age (days)", "l\u2093",         "lx",     True),
        ("Age-specific mortality (q\u2093)",   "Age (days)", "q\u2093",         "qx",     True),
        ("Life expectancy (e\u2093)",          "Age (days)", "Days remaining",  "ex",     True),
    ]

    axes = [fig.add_subplot(gs[r, c]) for r in range(2) for c in range(3)]

    for ax, (title, xlabel, ylabel, key, from_lt) in zip(axes, panel_cfg):
        for t, res in results.items():
            y = res["lt"][key].values if from_lt else res["sim"][key]
            x = res["lt"]["x"].values if from_lt else res["sim"]["days"]
            # filter out uninformative tail (qx / ex zeroes)
            if key in ("qx", "ex"):
                mask = res["sim"]["n_fem"] > 0
                x, y = x[mask], y[mask]
            ax.plot(x, y, color=res["p"]["color"], linewidth=1.8,
                    label=res["p"]["label"], alpha=0.9)
        ax.set_title(title, fontsize=9.5, fontweight="bold", pad=4)
        ax.set_xlabel(xlabel, fontsize=8)
        ax.set_ylabel(ylabel, fontsize=8)
        ax.tick_params(labelsize=7.5)
        ax.spines[["top", "right"]].set_visible(False)
        ax.grid(axis="y", alpha=0.25, linewidth=0.6)

    # Shared legend below figure
    handles = [plt.Line2D([0], [0], color=results[t]["p"]["color"],
                           linewidth=2, label=results[t]["p"]["label"])
               for t in results]
    fig.legend(handles=handles, loc="lower center", ncol=len(handles),
               fontsize=9, frameon=False,
               bbox_to_anchor=(0.5, 0.01))

    st.pyplot(fig, use_container_width=True)
    plt.close(fig)

    # PNG download
    buf = io.BytesIO()
    fig2 = plt.figure(figsize=(14, 8))  # regenerate for download
    fig2.patch.set_facecolor("white")
    gs2  = gridspec.GridSpec(2, 3, hspace=0.42, wspace=0.32,
                             left=0.07, right=0.97, top=0.93, bottom=0.09)
    axes2 = [fig2.add_subplot(gs2[r, c]) for r in range(2) for c in range(3)]
    for ax, (title, xlabel, ylabel, key, from_lt) in zip(axes2, panel_cfg):
        for t, res in results.items():
            y = res["lt"][key].values if from_lt else res["sim"][key]
            x = res["lt"]["x"].values if from_lt else res["sim"]["days"]
            if key in ("qx", "ex"):
                mask = res["sim"]["n_fem"] > 0
                x, y = x[mask], y[mask]
            ax.plot(x, y, color=res["p"]["color"], linewidth=1.8,
                    label=res["p"]["label"], alpha=0.9)
        ax.set_title(title, fontsize=9.5, fontweight="bold", pad=4)
        ax.set_xlabel(xlabel, fontsize=8); ax.set_ylabel(ylabel, fontsize=8)
        ax.tick_params(labelsize=7.5)
        ax.spines[["top", "right"]].set_visible(False)
        ax.grid(axis="y", alpha=0.25, linewidth=0.6)
    handles2 = [plt.Line2D([0], [0], color=results[t]["p"]["color"],
                            linewidth=2, label=results[t]["p"]["label"])
                for t in results]
    fig2.legend(handles=handles2, loc="lower center", ncol=len(handles2),
                fontsize=9, frameon=False, bbox_to_anchor=(0.5, 0.01))
    fig2.savefig(buf, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig2)
    buf.seek(0)
    st.download_button("⬇ Download figure (.png)", buf,
                       file_name="musca_comparison.png", mime="image/png")


# ── Tab 2: Life table ─────────────────────────────────────────────────────
with tab_lt:
    t_choice = st.selectbox(
        "Show life table for:",
        options=list(results.keys()),
        format_func=lambda t: LABELS[t],
    )
    lt_show = results[t_choice]["lt"].copy()
    lt_show.columns = [
        "x", "Nₓ", "mₓ", "lₓ", "dₓ", "qₓ", "Lₓ", "Tₓ", "eₓ (days)",
        "lₓ·mₓ", "x·lₓ·mₓ",
    ]
    # Round floats for display
    float_cols = [c for c in lt_show.columns if lt_show[c].dtype == float]
    lt_show[float_cols] = lt_show[float_cols].round(4)

    st.dataframe(
        lt_show,
        use_container_width=True,
        height=520,
        hide_index=True,
    )


# ── Tab 3: Demographics ────────────────────────────────────────────────────
with tab_dem:
    st.subheader("Comparative demographic parameters")

    PARAM_ROWS = [
        ("Net Reproductive Rate  R₀",             "R0"),
        ("R₀  daughters only  (1 : 1 sex ratio)", "R0_f"),
        ("Mean Generation Time  T  (days)",         "T_gen"),
        ("Intrinsic rate of increase  r  (/day)",   "r"),
        ("Finite rate of increase  λ",              "lam"),
        ("Population doubling time  Tₐ  (days)",   "Td"),
        ("Life expectancy at birth  e₀  (days)",   "e0"),
        ("Max observed longevity  (days)",          "max_life"),
        ("Total eggs collected",                    "tot_eggs"),
        ("Adults emerged from nursery",             "tot_emerged"),
        ("Immature survival  (%)",                  "imm_surv"),
    ]

    # Wide table for side-by-side comparison
    col_names = ["Parameter"] + [LABELS[t] for t in results]
    rows = []
    for label, key in PARAM_ROWS:
        row = [label] + [results[t]["dem"][key] for t in results]
        rows.append(row)
    dem_df = pd.DataFrame(rows, columns=col_names)
    st.dataframe(dem_df, use_container_width=True, hide_index=True, height=430)

    st.markdown("---")

    # Metric cards — one column per temperature
    cols = st.columns(len(results))
    highlight_keys = ["R0", "r", "lam", "e0", "T_gen"]
    metric_labels  = {
        "R0":    "Net Reproductive Rate  R₀",
        "r":     "Intrinsic rate  r  (/day)",
        "lam":   "Finite rate  λ",
        "e0":    "Life expectancy  e₀  (days)",
        "T_gen": "Mean Generation Time  (days)",
    }
    for col, t in zip(cols, results):
        dem = results[t]["dem"]
        col_hex = COLORS[t]
        with col:
            st.markdown(
                f"<div class='metric-card' style='--card-color:{col_hex}'>"
                f"<h4 style='color:{col_hex}'>{LABELS[t]}</h4>"
                "<table>" +
                "".join(
                    f"<tr><td>{metric_labels[k]}</td><td>{dem[k]}</td></tr>"
                    for k in highlight_keys
                ) +
                "</table></div>",
                unsafe_allow_html=True,
            )


# ── Tab 4: About ──────────────────────────────────────────────────────────
with tab_about:
    st.markdown("""
## About this simulation

This app models a laboratory cohort of *Musca domestica* (house fly) at up to
three temperatures simultaneously and constructs a formal **time-specific life table**.

### How the simulation works

| Step | What happens |
|------|-------------|
| **1 — Initialise** | 20 females + 20 males are created; each fly is assigned a random lifespan drawn from a truncated normal distribution |
| **2 — Daily loop** | Surviving females produce eggs each day; egg counts follow a Poisson distribution scaled by a Gaussian fecundity curve |
| **3 — Nursery** | Each egg batch develops through **egg → larva → pupa** stages with random stage durations and binomial mortality per day |
| **4 — Life table** | Standard columns computed: *N*ₓ, *l*ₓ, *d*ₓ, *q*ₓ, *L*ₓ, *T*ₓ, *e*ₓ, *m*ₓ |
| **5 — Demographics** | R₀, generation time *T*, intrinsic rate *r*, finite rate λ, doubling time *T*ₐ, and life expectancy *e*₀ |

### Stage durations — ADD method

Immature stage durations are derived using **Accumulated Degree-Days (ADD)**
with lower developmental threshold LDT = 12 °C:

| Stage | ADD (°C·days) |
|-------|--------------|
| Egg   | 9.68 |
| Larva | 72.57 |
| Pupa  | 81.91 |

At 25 °C, empirical values from Wang *et al.* (2018) are used directly.

### Key formulae

| Parameter | Formula |
|-----------|---------|
| R₀ | Σ(*l*ₓ · *m*ₓ) |
| Mean generation time T | Σ(x · *l*ₓ · *m*ₓ) / R₀ |
| Intrinsic rate r | ≈ ln(R₀) / T |
| Finite rate λ | eʳ |
| Doubling time Tₐ | ln 2 / r |

### References

Flint CM *et al.* (2025) *J Forensic Sci* 70(1): 169–178.

Ali MN *et al.* (2024) *Eur Chem Bull* 13(5): 10–25.

Krebs CJ (2014) *Ecology* 6th ed. Ch. 8.

Henderson PA (2016) *Practical Methods in Ecology* 4th ed. §12.4.
    """)


# ═════════════════════════════════════════════════════════════════════════════
# DOWNLOAD BUTTONS (sidebar, always visible)
# ═════════════════════════════════════════════════════════════════════════════
if results:
    with st.sidebar:
        st.markdown("---")
        st.subheader("⬇ Download results")

        # ── Excel workbook ──────────────────────────────────────────────────
        def build_excel(results: dict) -> bytes:
            wb = Workbook()

            # Helper styles
            hdr_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hdr_fill  = PatternFill("solid", fgColor="1D4ED8")
            data_font = Font(name="Calibri", size=10)
            thin      = Side(style="thin", color="CCCCCC")
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            centre    = Alignment(horizontal="center")

            def write_header(ws, headers, fill_color="1D4ED8"):
                fill = PatternFill("solid", fgColor=fill_color)
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c, value=h)
                    cell.font  = Font(name="Calibri", bold=True,
                                      size=10, color="FFFFFF")
                    cell.fill  = fill
                    cell.border = border
                    cell.alignment = centre

            # Comparison sheet
            ws0 = wb.active
            ws0.title = "Comparison"
            param_rows = [
                ("Temperature (°C)",                   "temp_C",     False),
                ("Net Reproductive Rate R0",            "R0",         True),
                ("R0 daughters only",                   "R0_f",       True),
                ("Mean Generation Time T (days)",       "T_gen",      True),
                ("Intrinsic rate r (/day)",             "r",          True),
                ("Finite rate lambda",                  "lam",        True),
                ("Doubling Time Td (days)",             "Td",         True),
                ("Life Expectancy e0 (days)",           "e0",         True),
                ("Max Longevity (days)",                "max_life",   False),
                ("Total Eggs Collected",                "tot_eggs",   False),
                ("Adults Emerged",                      "tot_emerged",False),
                ("Immature Survival (%)",               "imm_surv",   True),
            ]
            temp_labels = [results[t]["p"]["label"] for t in results]
            write_header(ws0, ["Parameter"] + temp_labels)
            for row_i, (name, key, do_round) in enumerate(param_rows, 2):
                ws0.cell(row=row_i, column=1, value=name).font = data_font
                for col_j, t in enumerate(results, 2):
                    src  = results[t]["p"] if key == "temp_C" else results[t]["dem"]
                    val  = src.get(key, src["p"].get(key, "—")) if isinstance(src, dict) else "—"
                    if do_round and isinstance(val, float):
                        val = round(val, 4)
                    cell = ws0.cell(row=row_i, column=col_j, value=val)
                    cell.font = data_font; cell.border = border
                    cell.alignment = centre
            ws0.column_dimensions["A"].width = 36
            for col in range(2, len(results) + 2):
                ws0.column_dimensions[get_column_letter(col)].width = 18

            # Per-temperature sheets
            fill_colors = {"15": "1565C0", "20": "2E7D32", "25": "B71C1C"}
            for t in results:
                r  = results[t]
                fc = fill_colors.get(t, "1D4ED8")

                # Population sheet
                ws_p = wb.create_sheet(f"{t}C_Population")
                pop_headers = ["Day", "Females alive", "Males alive", "Eggs collected"]
                write_header(ws_p, pop_headers, fill_color=fc)
                for i, d in enumerate(r["sim"]["days"]):
                    for c, v in enumerate([int(d), int(r["sim"]["n_fem"][i]),
                                           int(r["sim"]["n_mal"][i]),
                                           int(r["sim"]["eggs"][i])], 1):
                        cell = ws_p.cell(row=i+2, column=c, value=v)
                        cell.font = data_font; cell.border = border

                # Nursery sheet
                ws_n = wb.create_sheet(f"{t}C_Nursery")
                nur_headers = ["Egg Day", "Eggs Laid", "Egg Dur (d)", "Hatched",
                               "Larva Dur (d)", "Pupated", "Pupa Dur (d)",
                               "Emerged", "Emergence Day"]
                nur_keys    = ["egg_day","n_eggs","egg_dur","n_hatched",
                               "larva_dur","n_pupated","pupa_dur",
                               "n_emerged","emerge_day"]
                write_header(ws_n, nur_headers, fill_color=fc)
                for row_i, row in enumerate(r["sim"]["nursery"].itertuples(index=False), 2):
                    for c, k in enumerate(nur_keys, 1):
                        v = getattr(row, k)
                        v = None if (isinstance(v, float) and np.isnan(v)) else v
                        cell = ws_n.cell(row=row_i, column=c, value=v)
                        cell.font = data_font; cell.border = border

                # Life table sheet
                ws_l = wb.create_sheet(f"{t}C_LifeTable")
                lt_headers = ["x","Nx","mx","lx","dx","qx",
                               "Lx","Tx","ex (days)","lx*mx","x*lx*mx"]
                lt_keys    = ["x","Nx","mx","lx","dx","qx",
                               "Lx","Tx","ex","lx_mx","x_lx_mx"]
                write_header(ws_l, lt_headers, fill_color=fc)
                for row_i, row in enumerate(r["lt"].itertuples(index=False), 2):
                    for c, k in enumerate(lt_keys, 1):
                        v = getattr(row, k)
                        if isinstance(v, float):
                            v = round(v, 5)
                        cell = ws_l.cell(row=row_i, column=c, value=v)
                        cell.font = data_font; cell.border = border

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        excel_bytes = build_excel(results)
        st.download_button(
            "📥 Excel workbook (.xlsx)",
            data=excel_bytes,
            file_name="musca_simulation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # ── Text summary ────────────────────────────────────────────────────
        def build_txt(results: dict) -> str:
            lines = [
                "=" * 68,
                "  MUSCA DOMESTICA THREE-TEMPERATURE COHORT SIMULATION",
                "  Comparative Life Table Demographic Parameters",
                "=" * 68, "",
                "  Methods:",
                "    Krebs CJ (2014) Ecology 6th ed. Ch. 8",
                "    Henderson PA (2016) Ecological Methods 4th ed. §12.4",
                "    Stage durations: ADD method (LDT = 12°C)",
                "    Flint et al. (2025) J Forensic Sci",
                "    Ali et al. (2024) Eur Chem Bull", "",
            ]
            for t, res in results.items():
                d = res["dem"]; p = res["p"]
                eff = p["temp_C"] - 12
                lines += [
                    f"  {p['label']}  (effective temp above LDT: {eff}°C)",
                    "-" * 50,
                    f"    Net Reproductive Rate  R0          : {d['R0']:>10.3f}  eggs/female/lifetime",
                    f"    R0 daughters only (1:1 ratio)      : {d['R0_f']:>10.3f}  daughters/female",
                    f"    Mean Generation Time  T            : {d['T_gen']:>10.2f}  days",
                    f"    Intrinsic rate  r                  : {d['r']:>10.4f}  per day",
                    f"    Finite rate  lambda                : {d['lam']:>10.4f}",
                    f"    Population doubling time  Td       : {d['Td']:>10.2f}  days",
                    f"    Life expectancy at birth  e0       : {d['e0']:>10.2f}  days",
                    f"    Max observed longevity             : {d['max_life']:>10d}  days",
                    f"    Total eggs collected               : {d['tot_eggs']:>10d}",
                    f"    Adults emerged from nursery        : {d['tot_emerged']:>10d}",
                    f"    Immature survival                  : {d['imm_surv']:>9.1f}%",
                    "",
                ]
            lines.append("=" * 68)
            return "\n".join(lines)

        txt_content = build_txt(results)
        st.download_button(
            "📄 Parameter summary (.txt)",
            data=txt_content,
            file_name="musca_parameters.txt",
            mime="text/plain",
            use_container_width=True,
        )
